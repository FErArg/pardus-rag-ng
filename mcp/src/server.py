#!/usr/bin/env python3
"""
PardusDB MCP Server

Model Context Protocol server for PardusDB vector database.
Enables AI agents to perform vector similarity search, manage vector data,
import documents, and perform health checks.
"""

import asyncio
import csv
import hashlib
import json
import os
import shutil
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
import re
from typing import Any, List, Optional

try:
    from .model_context import (
        get_context_window_for_model,
        detect_provider,
        DEFAULT_TOKENS_PER_CHUNK as MCP_TOKENS_PER_CHUNK,
    )
except ImportError:
    from model_context import (
        get_context_window_for_model,
        detect_provider,
        DEFAULT_TOKENS_PER_CHUNK as MCP_TOKENS_PER_CHUNK,
    )

try:
    from mcp.server import Server
    from mcp.server.stdio import stdio_server
    from mcp.types import Tool, TextContent
except ImportError as exc:
    is_macos = sys.platform == "darwin"
    if is_macos:
        print("Error: mcp package not found.", file=sys.stderr)
        print("On macOS, run this script via ~/.pardus/mcp/run_mcp.sh, not directly with python3.", file=sys.stderr)
        print("(The installer creates a virtual environment that contains the mcp package.)", file=sys.stderr)
    else:
        print("Error: mcp package not found. Install with: pip install mcp", file=sys.stderr)
    sys.exit(1)


MAX_FILE_SIZE_MB = 50
DEFAULT_VECTOR_DIM = 384
EMBEDDER_MODEL = "all-MiniLM-L6-v2"
TMP_DIR = Path("./tmp")
STATS_FILE = Path.home() / ".pardus" / "mcp_stats.json"
DEFAULT_MODEL = "claude-3-5-sonnet-20241022"
DEFAULT_CONTEXT_WINDOW = 1000000
DEFAULT_TOKENS_PER_CHUNK = 300


def _ensure_tmp_dir() -> Path:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    return TMP_DIR


def _create_tmp_uuid_dir() -> Path:
    _ensure_tmp_dir()
    uid = uuid.uuid4().hex
    tmp_uuid_dir = TMP_DIR / uid
    tmp_uuid_dir.mkdir(parents=True, exist_ok=True)
    return tmp_uuid_dir


def _convert_to_markdown(file_path: str, output_dir: Path) -> tuple[Path, str]:
    path = Path(file_path)
    ext = path.suffix.lower()
    parser = PARSERS.get(ext)
    if not parser:
        raise ValueError(f"Unsupported file type: {ext}")
    parsed = parser(str(path))
    content = parsed.get("content", "")
    md_path = output_dir / f"{path.stem}.md"
    md_path.write_text(content, encoding="utf-8")
    return md_path, ext


def _cleanup_tmp_dir(tmp_uuid_dir: Path):
    if tmp_uuid_dir.exists():
        shutil.rmtree(tmp_uuid_dir, ignore_errors=True)


# ==================== Token Tracking ====================

def _load_stats() -> dict:
    """Load stats from JSON file."""
    if STATS_FILE.exists():
        try:
            with open(STATS_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "config": {
            "model": DEFAULT_MODEL,
            "context_window": DEFAULT_CONTEXT_WINDOW,
            "provider": "anthropic"
        },
        "session": {
            "start": datetime.now(timezone.utc).isoformat(),
            "queries": 0,
            "tokens_sent": 0,
            "tokens_if_full": 0,
            "chunks_returned": 0
        },
        "total": {
            "queries": 0,
            "tokens_sent": 0,
            "tokens_if_full": 0,
            "chunks_returned": 0
        }
    }


def _save_stats(stats: dict) -> None:
    """Save stats to JSON file."""
    STATS_FILE.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(STATS_FILE, "w") as f:
            json.dump(stats, f, indent=2)
    except Exception:
        pass


def update_token_stats(chunks_returned: int, tokens_sent: int) -> dict:
    """Update token stats after a search query."""
    stats = _load_stats()

    # Calculate what would have been sent without MCP
    tokens_if_full = tokens_sent * 20  # Rough estimate: full docs would be ~20x larger

    # Update session
    stats["session"]["queries"] += 1
    stats["session"]["tokens_sent"] += tokens_sent
    stats["session"]["tokens_if_full"] += tokens_if_full
    stats["session"]["chunks_returned"] += chunks_returned

    # Update total
    stats["total"]["queries"] += 1
    stats["total"]["tokens_sent"] += tokens_sent
    stats["total"]["tokens_if_full"] += tokens_if_full
    stats["total"]["chunks_returned"] += chunks_returned

    _save_stats(stats)
    return stats


def get_token_stats() -> dict:
    """Get current token stats with calculated savings."""
    stats = _load_stats()

    tokens_sent = stats["session"]["tokens_sent"]
    tokens_if_full = stats["session"]["tokens_if_full"]
    savings = tokens_if_full - tokens_sent if tokens_if_full > 0 else 0
    savings_pct = (savings / tokens_if_full * 100) if tokens_if_full > 0 else 0

    total_tokens_sent = stats["total"]["tokens_sent"]
    total_tokens_if_full = stats["total"]["tokens_if_full"]
    total_savings = total_tokens_if_full - total_tokens_sent if total_tokens_if_full > 0 else 0
    total_savings_pct = (total_savings / total_tokens_if_full * 100) if total_tokens_if_full > 0 else 0

    return {
        "config": stats["config"],
        "session": {
            "start": stats["session"]["start"],
            "queries": stats["session"]["queries"],
            "tokens_sent": tokens_sent,
            "tokens_if_full": tokens_if_full,
            "savings": savings,
            "savings_percent": round(savings_pct, 1),
            "chunks_returned": stats["session"]["chunks_returned"]
        },
        "total": {
            "queries": stats["total"]["queries"],
            "tokens_sent": total_tokens_sent,
            "tokens_if_full": total_tokens_if_full,
            "savings": total_savings,
            "savings_percent": round(total_savings_pct, 1),
            "chunks_returned": stats["total"]["chunks_returned"]
        }
    }


def set_current_model(model: str) -> dict:
    """Set the current model and update config."""
    stats = _load_stats()
    context_window = get_context_window_for_model(model)
    provider = detect_provider(model)

    stats["config"] = {
        "model": model,
        "context_window": context_window,
        "provider": provider
    }
    _save_stats(stats)
    return {"model": model, "context_window": context_window, "provider": provider}


def reset_session_stats() -> dict:
    """Reset session stats (keep total)."""
    stats = _load_stats()
    stats["session"] = {
        "start": datetime.now(timezone.utc).isoformat(),
        "queries": 0,
        "tokens_sent": 0,
        "tokens_if_full": 0,
        "chunks_returned": 0
    }
    _save_stats(stats)
    return {"status": "Session stats reset"}


# ==================== Optional Dependencies ====================

_embedder_instance = None
HAS_EMBEDDER = False

try:
    from sentence_transformers import SentenceTransformer
    HAS_EMBEDDER = True
except ImportError:
    pass

def get_embedder():
    global _embedder_instance
    if _embedder_instance is None and HAS_EMBEDDER:
        try:
            from sentence_transformers import SentenceTransformer
            _embedder_instance = SentenceTransformer(EMBEDDER_MODEL)
        except Exception as e:
            print(f"[embedder] failed to load model: {e}", file=sys.stderr)
    return _embedder_instance

try:
    import pypdf
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    import xlrd
    XLS_AVAILABLE = True
except ImportError:
    XLS_AVAILABLE = False


# ==================== Types ====================

class PardusDBClient:
    def __init__(self) -> None:
        self.db_path: Optional[str] = None
        self.current_table: Optional[str] = None
        self._discover_database()

    def _discover_database(self) -> None:
        """Check for database.pardus in CWD. If found, open and verify integrity."""
        cwd_db = Path.cwd() / "database.pardus"
        if cwd_db.exists():
            self.db_path = str(cwd_db)
            self._verify_integrity()

    def _verify_integrity(self) -> None:
        """Run basic integrity check: access + SHOW TABLES."""
        result = self.execute("SHOW TABLES;")
        if "Error" in result:
            raise ConnectionError(f"Database integrity check failed: {result}")

    def execute(self, command: str) -> str:
        import subprocess
        if self.db_path is None:
            cwd_db = Path.cwd() / "database.pardus"
            if cwd_db.exists():
                self.db_path = str(cwd_db)
                self._verify_integrity()
            else:
                self.db_path = str(cwd_db)
                create_result = self._create_database(str(cwd_db))
                if "Error" in create_result:
                    return create_result
        db_arg = [self.db_path] if self.db_path else []
        try:
            proc = subprocess.run(
                ["pardusdb", *db_arg],
                input=f"{command}\nsave\nquit\n".encode(),
                capture_output=True,
                timeout=30,
            )
            output = (proc.stdout + proc.stderr).decode()
            if proc.returncode != 0:
                return f"Error (exit {proc.returncode}): {output}"
            return output
        except subprocess.TimeoutExpired:
            return "Error: Query timed out"
        except FileNotFoundError:
            return "Error: pardusdb binary not found in PATH"

    def _create_database(self, path: str) -> str:
        """Create a new database file."""
        import subprocess
        try:
            proc = subprocess.run(
                ["pardusdb", path],
                input=b".save\nquit\n",
                capture_output=True,
                timeout=30,
            )
            output = (proc.stdout + proc.stderr).decode()
            if proc.returncode != 0:
                return f"Error (exit {proc.returncode}): {output}"
            return output
        except Exception as e:
            return f"Error creating database: {e}"

    def set_db_path(self, db_path: Optional[str]) -> None:
        self.db_path = db_path

    def get_db_path(self) -> Optional[str]:
        return self.db_path

    def set_current_table(self, table_name: Optional[str]) -> None:
        self.current_table = table_name

    def get_current_table(self) -> Optional[str]:
        return self.current_table


db_client = PardusDBClient()


# ==================== SQL Escaping Helpers ====================

def sql_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "''").replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t")


def sql_escape_path(path: str) -> str:
    return sql_escape(str(Path(path).resolve()))


def sql_safe_identifier(name: str) -> str:
    if not name:
        raise ValueError("Empty identifier not allowed")
    if not all(c.isalnum() or c == '_' for c in name):
        raise ValueError(f"Invalid identifier: {name}")
    return name


# ==================== Helper Functions ====================

def file_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def generate_embedding(text: str, dim: int) -> list[float]:
    if not text or not HAS_EMBEDDER:
        return [0.0] * dim
    embedder = get_embedder()
    if embedder is None:
        return [0.0] * dim
    try:
        vec = embedder.encode(text, convert_to_numpy=True, normalize_embeddings=True)
        if vec is None:
            return [0.0] * dim
        vec = vec.tolist() if hasattr(vec, 'tolist') else list(vec)
        if len(vec) != dim:
            return [0.0] * dim
        return vec
    except Exception as e:
        print(f"[embedder] error: {e}", file=sys.stderr)
        return [0.0] * dim


def get_table_schema(table: str) -> dict[str, Any]:
    schema = {
        "exists": False,
        "columns": [],
        "row_count": 0,
        "vector_dim": None,
    }
    try:
        safe_table = sql_safe_identifier(table)
        result = db_client.execute("SHOW TABLES")
        if table in result:
            schema["exists"] = True
        count_result = db_client.execute(f"SELECT COUNT(*) FROM {safe_table}")
        schema["row_count"] = parse_count_from_result(count_result)
    except Exception:
        pass
    return schema


def ensure_import_table(table: str, dim: int) -> None:
    safe_table = sql_safe_identifier(table)
    db_client.execute(f"CREATE TABLE IF NOT EXISTS {safe_table} (embedding VECTOR({dim}), filename TEXT, content TEXT, page INT, file_type TEXT, parent_doc_id INT, doc_path TEXT, chunk_index INT, total_chunks INT, title TEXT)")
    db_client.execute("CREATE TABLE IF NOT EXISTS __import_log__ (id INTEGER PRIMARY KEY, table_name TEXT, doc_path TEXT, filename TEXT, file_size INT, file_hash TEXT, content_hash TEXT, imported_at TEXT, total_parents INT, total_children INT, status TEXT)")


def log_import(
    table: str,
    doc_path: str,
    filename: str,
    file_size: int,
    file_hash: str,
    content_hash: str,
    total_parents: int,
    total_children: int,
    status: str,
) -> None:
    now = datetime.now(timezone.utc).isoformat()
    table_esc = sql_escape(table)
    doc_path_esc = sql_escape(doc_path)
    filename_esc = sql_escape(filename)
    sql = f"""INSERT INTO __import_log__
        (table_name, doc_path, filename, file_size, file_hash, content_hash, imported_at, total_parents, total_children, status)
        VALUES ('{table_esc}', '{doc_path_esc}', '{filename_esc}', {file_size}, '{file_hash}', '{content_hash}', '{now}', {total_parents}, {total_children}, '{status}')"""
    try:
        db_client.execute(sql)
    except Exception:
        pass


def is_already_imported(table: str, file_hash: str, content_hash: str) -> bool:
    table_esc = sql_escape(table)
    sql = f"SELECT COUNT(*) FROM __import_log__ WHERE table_name = '{table_esc}' AND (file_hash = '{file_hash}' OR content_hash = '{content_hash}')"
    try:
        result = db_client.execute(sql)
        return parse_count_from_result(result) > 0
    except Exception:
        pass
    return False


def parse_id_from_result(result: str) -> Optional[int]:
    m = re.search(r"id=(\d+)", result)
    if not m:
        return None
    return int(m.group(1))


def parse_count_from_result(result: str) -> int:
    m = re.search(r"Count:\s*(\d+)", result)
    if not m:
        return 0
    return int(m.group(1))


def smart_chunk(text: str, target_chars: int = 500, overlap: int = 50) -> List[str]:
    """Split text into ~target_chars chunks at sentence boundaries with overlap."""
    if not text or not text.strip():
        return []
    sentences = re.split(r'(?<=[.!?])\s+', text.strip())
    chunks, current = [], ""
    for sent in sentences:
        if len(current) + len(sent) <= target_chars:
            current += sent + " "
        else:
            if current.strip():
                chunks.append(current.strip())
            current = (current[-overlap:] + sent + " ") if overlap > 0 and current else sent + " "
    if current.strip():
        chunks.append(current.strip())
    return chunks


def compute_chunk_hash(chunk: str) -> str:
    """Compute a short hash for deduplication."""
    return hashlib.sha256(chunk.encode()).hexdigest()[:16]


# ==================== File Parsers ====================

def parse_txt(path: str) -> dict[str, Any]:
    content = Path(path).read_text(encoding="utf-8", errors="replace")
    return {
        "title": Path(path).stem,
        "content": content,
        "pages": [{"content": content, "page": 0}],
    }


def parse_md(path: str) -> dict[str, Any]:
    content = Path(path).read_text(encoding="utf-8", errors="replace")
    title = Path(path).stem
    return {
        "title": title,
        "content": content,
        "pages": [{"content": content, "page": 0}],
    }


def parse_csv(path: str) -> dict[str, Any]:
    rows = []
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                rows.append({"content": json.dumps(row), "page": i + 1})
    except Exception:
        content = Path(path).read_text(encoding="utf-8", errors="replace")
        return {
            "title": Path(path).stem,
            "content": content,
            "pages": [{"content": content, "page": 0}],
        }
    full_content = Path(path).read_text(encoding="utf-8", errors="replace")
    return {
        "title": Path(path).stem,
        "content": full_content,
        "pages": rows if rows else [{"content": full_content, "page": 0}],
    }


def parse_pdf(path: str) -> dict[str, Any]:
    if not PDF_AVAILABLE:
        raise ImportError("pypdf not installed. Install with: pip install pypdf")
    reader = pypdf.PdfReader(path)
    title = Path(path).stem
    try:
        if reader.metadata and reader.metadata.get("/Title"):
            title = reader.metadata.get("/Title", title)
    except Exception:
        pass
    pages = []
    for i, page in enumerate(reader.pages):
        try:
            text = page.extract_text()
        except Exception:
            text = ""
        if text.strip():
            pages.append({"content": text, "page": i + 1})
    full_content = "\n\n".join(p["content"] for p in pages)
    return {
        "title": title,
        "content": full_content,
        "pages": pages,
    }


def parse_docx(path: str) -> dict[str, Any]:
    if not DOCX_AVAILABLE:
        raise ImportError("python-docx not installed. Install with: pip install python-docx")
    doc = docx.Document(path)
    title = Path(path).stem
    try:
        if doc.core_properties.title:
            title = doc.core_properties.title
    except Exception:
        pass
    paragraphs = []
    for i, para in enumerate(doc.paragraphs):
        text = para.text.strip()
        if text:
            paragraphs.append({"content": text, "page": i + 1})
    full_content = "\n\n".join(p["content"] for p in paragraphs)
    return {
        "title": title,
        "content": full_content,
        "pages": paragraphs if paragraphs else [{"content": full_content, "page": 0}],
    }


def parse_xlsx(path: str) -> dict[str, Any]:
    if not XLSX_AVAILABLE:
        raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    title = Path(path).stem
    if wb.sheetnames:
        title = wb.sheetnames[0]
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_data = {str(cell.column_letter): str(cell.value) if cell.value is not None else "" for cell in row}
            if any(v for v in row_data.values()):
                rows.append({"content": json.dumps(row_data), "page": i})
    full_content = "\n".join(r["content"] for r in rows)
    return {
        "title": title,
        "content": full_content,
        "pages": rows if rows else [{"content": "", "page": 0}],
    }


def parse_json(path: str) -> dict[str, Any]:
    content = Path(path).read_text(encoding="utf-8", errors="replace")
    try:
        data = json.loads(content)
    except Exception:
        return {
            "title": Path(path).stem,
            "content": content,
            "pages": [{"content": content, "page": 0}],
        }
    pages = []
    if isinstance(data, list):
        for i, item in enumerate(data):
            if isinstance(item, dict):
                text = item.get("text", json.dumps(item))
                pages.append({"content": text, "page": i + 1})
            else:
                pages.append({"content": str(item), "page": i + 1})
    elif isinstance(data, dict):
        text = data.get("text", data.get("content", json.dumps(data)))
        pages.append({"content": text, "page": 0})
    full_content = "\n".join(p["content"] for p in pages)
    return {
        "title": Path(path).stem,
        "content": full_content,
        "pages": pages if pages else [{"content": content, "page": 0}],
    }


def parse_jsonl(path: str) -> dict[str, Any]:
    lines = Path(path).read_text(encoding="utf-8", errors="replace").strip().split("\n")
    pages = []
    for i, line in enumerate(lines, start=1):
        try:
            item = json.loads(line)
        except Exception:
            item = {"text": line}
        if isinstance(item, dict):
            text = item.get("text", item.get("content", json.dumps(item)))
        else:
            text = str(item)
        pages.append({"content": text, "page": i})
    full_content = "\n".join(p["content"] for p in pages)
    return {
        "title": Path(path).stem,
        "content": full_content,
        "pages": pages,
    }


def parse_xls(path: str) -> dict[str, Any]:
    if not XLS_AVAILABLE:
        raise ImportError("xlrd not installed. Install with: pip install xlrd")
    wb = xlrd.open_workbook(path)
    title = Path(path).stem
    sheet = wb.sheet_by_index(0)
    rows = []
    for i in range(sheet.nrows):
        row_data = {}
        for j in range(sheet.ncols):
            cell = sheet.cell(i, j)
            row_data[str(j)] = str(cell.value) if cell.value else ""
        if any(v for v in row_data.values()):
            rows.append({"content": json.dumps(row_data), "page": i + 1})
    full_content = "\n".join(r["content"] for r in rows)
    return {
        "title": title,
        "content": full_content,
        "pages": rows if rows else [{"content": "", "page": 0}],
    }


PARSERS = {
    ".txt": parse_txt,
    ".md": parse_md,
    ".csv": parse_csv,
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    ".xls": parse_xls,
    ".json": parse_json,
    ".jsonl": parse_jsonl,
}

SUPPORTED_EXTENSIONS = list(PARSERS.keys())


# ==================== Tool Handlers ====================

async def handle_create_database(args: dict[str, Any]) -> dict[str, Any]:
    db_path = args.get("path")
    if not db_path:
        return {"content": [{"type": "text", "text": "Error: Database path is required"}], "isError": True}
    try:
        parent = Path(db_path).parent
        if parent and not parent.exists():
            parent.mkdir(parents=True, exist_ok=True)
        db_client.set_db_path(db_path)
        if Path(db_path).exists():
            db_client.execute(f".open {db_path}")
            return {"content": [{"type": "text", "text": f"Database opened (already exists): {db_path}"}]}
        else:
            db_client.execute(f".create {db_path}")
            return {"content": [{"type": "text", "text": f"Database created successfully at: {db_path}"}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error creating database: {e}"}], "isError": True}


async def handle_open_database(args: dict[str, Any]) -> dict[str, Any]:
    db_path = args.get("path")
    if not db_path:
        return {"content": [{"type": "text", "text": "Error: Database path is required"}], "isError": True}
    if not Path(db_path).exists():
        return {"content": [{"type": "text", "text": f"Error: Database file not found: {db_path}"}], "isError": True}
    try:
        db_client.set_db_path(db_path)
        db_client.execute(f".open {db_path}")
        return {"content": [{"type": "text", "text": f"Database opened successfully: {db_path}"}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error opening database: {e}"}], "isError": True}


async def handle_create_table(args: dict[str, Any]) -> dict[str, Any]:
    name = args.get("name")
    vector_dim = args.get("vector_dim")
    metadata_schema = args.get("metadata_schema")
    if not name or not vector_dim:
        return {"content": [{"type": "text", "text": "Error: Table name and vector_dim are required"}], "isError": True}
    try:
        safe_name = sql_safe_identifier(name)
        columns = [f"embedding VECTOR({vector_dim})", "filename TEXT", "content TEXT", "page INT", "file_type TEXT", "parent_doc_id INT", "doc_path TEXT", "chunk_index INT", "total_chunks INT", "title TEXT"]
        type_map = {
            "str": "TEXT", "string": "TEXT",
            "int": "INTEGER", "integer": "INTEGER",
            "float": "FLOAT", "bool": "BOOLEAN", "text": "TEXT",
        }
        if metadata_schema:
            for col_name, col_type in metadata_schema.items():
                safe_col = sql_safe_identifier(col_name)
                sql_type = type_map.get(col_type.lower(), col_type.upper())
                columns.append(f"{safe_col} {sql_type}")
        sql = f"CREATE TABLE {safe_name} ({', '.join(columns)})"
        db_client.execute(sql)
        db_client.set_current_table(name)
        return {"content": [{"type": "text", "text": f"Table '{name}' created successfully with {vector_dim}-dimensional vectors.\n\nSQL: {sql}"}]}
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid input: {e}"}], "isError": True}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error creating table: {e}"}], "isError": True}


async def handle_insert_vector(args: dict[str, Any]) -> dict[str, Any]:
    vector = args.get("vector")
    metadata = args.get("metadata")
    table = args.get("table") or db_client.get_current_table()
    if not vector or not isinstance(vector, list):
        return {"content": [{"type": "text", "text": "Error: Vector array is required"}], "isError": True}
    if not table:
        return {"content": [{"type": "text", "text": "Error: No table specified. Use 'use_table' first or provide 'table' parameter."}], "isError": True}
    try:
        safe_table = sql_safe_identifier(table)
        columns = ["embedding"]
        values = [f"[{', '.join(str(x) for x in vector)}]"]
        if metadata:
            for key, val in metadata.items():
                safe_key = sql_safe_identifier(key)
                columns.append(safe_key)
                if isinstance(val, str):
                    values.append(f"'{sql_escape(val)}'")
                elif isinstance(val, bool):
                    values.append("true" if val else "false")
                else:
                    values.append(str(val))
        sql = f"INSERT INTO {safe_table} ({', '.join(columns)}) VALUES ({', '.join(values)})"
        result = db_client.execute(sql)
        id_match = parse_id_from_result(result) or "unknown"
        return {"content": [{"type": "text", "text": f"Vector inserted successfully with ID: {id_match}"}]}
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid input: {e}"}], "isError": True}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error inserting vector: {e}"}], "isError": True}


async def handle_batch_insert(args: dict[str, Any]) -> dict[str, Any]:
    vectors = args.get("vectors")
    metadata_list = args.get("metadata_list")
    table = args.get("table") or db_client.get_current_table()
    if not vectors or not isinstance(vectors, list):
        return {"content": [{"type": "text", "text": "Error: Vectors array is required"}], "isError": True}
    if not table:
        return {"content": [{"type": "text", "text": "Error: No table specified"}], "isError": True}
    try:
        safe_table = sql_safe_identifier(table)
        results = []
        for i, vector in enumerate(vectors):
            metadata = metadata_list[i] if metadata_list else None
            columns = ["embedding"]
            values = [f"[{', '.join(str(x) for x in vector)}]"]
            if metadata:
                for key, val in metadata.items():
                    safe_key = sql_safe_identifier(key)
                    columns.append(safe_key)
                    if isinstance(val, str):
                        values.append(f"'{sql_escape(val)}'")
                    elif isinstance(val, bool):
                        values.append("true" if val else "false")
                    else:
                        values.append(str(val))
            sql = f"INSERT INTO {safe_table} ({', '.join(columns)}) VALUES ({', '.join(values)})"
            result = db_client.execute(sql)
            vid = parse_id_from_result(result)
            if vid:
                results.append(str(vid))
        return {"content": [{"type": "text", "text": f"Batch insert completed. Inserted {len(results)} vectors with IDs: {', '.join(results)}"}]}
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid input: {e}"}], "isError": True}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error during batch insert: {e}"}], "isError": True}


async def handle_search_similar(args: dict[str, Any]) -> dict[str, Any]:
    query_vector = args.get("query_vector")
    k = args.get("k", 10)
    table = args.get("table") or db_client.get_current_table()
    if not query_vector or not isinstance(query_vector, list):
        return {"content": [{"type": "text", "text": "Error: query_vector array is required"}], "isError": True}
    if not table:
        return {"content": [{"type": "text", "text": "Error: No table specified"}], "isError": True}
    try:
        safe_table = sql_safe_identifier(table)
        vector_str = f"[{', '.join(str(x) for x in query_vector)}]"
        sql = f"SELECT * FROM {safe_table} WHERE embedding SIMILARITY {vector_str} LIMIT {k}"
        result = db_client.execute(sql)
        return {"content": [{"type": "text", "text": f"Search Results:\n\n{result}"}]}
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid input: {e}"}], "isError": True}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error searching: {e}"}], "isError": True}


async def handle_search_text(args: dict[str, Any]) -> dict[str, Any]:
    query = args.get("query")
    k = args.get("k", 10)
    table = args.get("table") or db_client.get_current_table()
    vector_dim = args.get("vector_dim", DEFAULT_VECTOR_DIM)

    if not query or not isinstance(query, str):
        return {"content": [{"type": "text", "text": "Error: query string is required"}], "isError": True}
    if not table:
        return {"content": [{"type": "text", "text": "Error: No table specified"}], "isError": True}

    try:
        query_vector = generate_embedding(query, vector_dim)
        if all(v == 0.0 for v in query_vector) and HAS_EMBEDDER:
            return {"content": [{"type": "text", "text": "Error: Failed to generate embedding for query"}], "isError": True}

        safe_table = sql_safe_identifier(table)
        vector_str = f"[{', '.join(str(x) for x in query_vector)}]"
        sql = f"SELECT * FROM {safe_table} WHERE embedding SIMILARITY {vector_str} LIMIT {k}"
        result = db_client.execute(sql)

        # Estimate tokens
        tokens_sent = k * MCP_TOKENS_PER_CHUNK
        stats = update_token_stats(chunks_returned=k, tokens_sent=tokens_sent)

        return {
            "content": [{"type": "text", "text": f"Search Results for '{query}':\n\n{result}"}],
            "stats": {
                "chunks_returned": k,
                "tokens_sent": tokens_sent,
                "session_tokens_total": stats["session"]["tokens_sent"],
                "session_savings_percent": stats["session"]["savings_percent"],
            }
        }
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid input: {e}"}], "isError": True}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error searching: {e}"}], "isError": True}


async def handle_execute_sql(args: dict[str, Any]) -> dict[str, Any]:
    sql = args.get("sql")
    if not sql:
        return {"content": [{"type": "text", "text": "Error: SQL query is required"}], "isError": True}
    try:
        result = db_client.execute(sql)
        return {"content": [{"type": "text", "text": f"Query Result:\n\n{result}"}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error executing SQL: {e}"}], "isError": True}


async def handle_list_tables() -> dict[str, Any]:
    try:
        result = db_client.execute("SHOW TABLES")
        return {"content": [{"type": "text", "text": f"Tables:\n\n{result}"}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error listing tables: {e}"}], "isError": True}


async def handle_use_table(args: dict[str, Any]) -> dict[str, Any]:
    table = args.get("table")
    if not table:
        return {"content": [{"type": "text", "text": "Error: Table name is required"}], "isError": True}
    db_client.set_current_table(table)
    return {"content": [{"type": "text", "text": f"Now using table: {table}"}]}


async def handle_get_status() -> dict[str, Any]:
    db_path = db_client.get_db_path()
    current_table = db_client.get_current_table()
    status = "PardusDB Status:\n\n"
    status += f"Database: {db_path or 'Not opened (in-memory)'}\n"
    status += f"Current Table: {current_table or 'None selected'}\n"
    if db_path and Path(db_path).exists():
        size = os.path.getsize(db_path)
        status += f"Database Size: {size / 1024:.2f} KB\n"
    return {"content": [{"type": "text", "text": status}]}


# ==================== Import Tool ====================

async def handle_import_text(args: dict[str, Any]) -> dict[str, Any]:
    dir_path = args.get("dir_path")
    table = args.get("table")
    file_patterns = args.get("file_patterns", SUPPORTED_EXTENSIONS)
    recursive = args.get("recursive", True)
    max_file_size_mb = args.get("max_file_size_mb", MAX_FILE_SIZE_MB)
    vector_dim = args.get("vector_dim", DEFAULT_VECTOR_DIM)
    tmp_uuid_dir = None

    if not dir_path or not table:
        return {"content": [{"type": "text", "text": "Error: dir_path and table are required"}], "isError": True}

    try:
        safe_table = sql_safe_identifier(table)
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid table name: {e}"}], "isError": True}

    if not Path(dir_path).exists():
        return {"content": [{"type": "text", "text": f"Error: Directory not found: {dir_path}"}], "isError": True}

    ensure_import_table(safe_table, vector_dim)

    all_files = []
    base_path = Path(dir_path).resolve()
    pattern_set = set(file_patterns) if file_patterns else set(SUPPORTED_EXTENSIONS)

    try:
        if recursive:
            for ext in pattern_set:
                all_files.extend(base_path.rglob(f"*{ext}"))
        else:
            for ext in pattern_set:
                all_files.extend(base_path.glob(f"*{ext}"))
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error scanning directory: {e}"}], "isError": True}

    all_files = sorted(set(all_files))
    total_files = len(all_files)

    if total_files == 0:
        return {"content": [{"type": "text", "text": f"No supported files found in {dir_path} with patterns {file_patterns}"}]}

    try:
        tmp_uuid_dir = _create_tmp_uuid_dir()
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error creating tmp directory: {e}"}], "isError": True}

    stats = {"imported": 0, "skipped": 0, "errors": 0}
    error_details = []
    imported_files = []

    for i, file_path in enumerate(all_files):
        try:
            resolved_path = file_path.resolve()
            if not str(resolved_path).startswith(str(base_path) + os.sep):
                stats["skipped"] += 1
                error_details.append(f"Skipped (symlink outside base dir): {file_path.name}")
                continue
        except Exception:
            stats["skipped"] += 1
            error_details.append(f"Skipped (cannot resolve path): {file_path.name}")
            continue

        fpath = str(file_path)
        fsize = 0
        try:
            fsize = file_path.stat().st_size
        except Exception:
            pass

        if max_file_size_mb and fsize > max_file_size_mb * 1024 * 1024:
            stats["skipped"] += 1
            error_details.append(f"Skipped (too large): {file_path.name} ({fsize / 1024 / 1024:.1f} MB)")
            continue

        ext = file_path.suffix.lower()
        parser = PARSERS.get(ext)

        if not parser:
            stats["skipped"] += 1
            continue

        try:
            fhash = file_hash(fpath)
        except Exception:
            fhash = "unknown"

        try:
            md_path, _ = _convert_to_markdown(fpath, tmp_uuid_dir)
            parsed = parse_md(str(md_path))
        except ValueError as e:
            stats["errors"] += 1
            error_details.append(f"{file_path.name}: {str(e)}")
            continue
        except ImportError as e:
            stats["errors"] += 1
            error_details.append(f"{file_path.name}: {str(e)}")
            continue
        except Exception as e:
            stats["errors"] += 1
            error_details.append(f"{file_path.name}: {str(e)}")
            continue

        title = parsed.get("title", file_path.stem)
        content = parsed.get("content", "")
        pages = parsed.get("pages", [])
        total_chunks = len(pages) if pages else 1

        try:
            chash = hashlib.sha256(content.encode()).hexdigest() if content else "empty"
        except Exception:
            chash = "unknown"

        if is_already_imported(safe_table, fhash, chash):
            stats["skipped"] += 1
            continue

        embedder_failed = False
        try:
            page_texts = [p.get("content", "") for p in pages]
            if page_texts and HAS_EMBEDDER:
                embedder = get_embedder()
                if embedder:
                    try:
                        vecs = embedder.encode(page_texts, convert_to_numpy=True, normalize_embeddings=True)
                        vecs = vecs.tolist() if hasattr(vecs, 'tolist') else list(vecs)
                        if len(vecs) != len(page_texts) or (vecs and len(vecs[0]) != vector_dim):
                            vecs = [[0.0] * vector_dim] * len(page_texts)
                            embedder_failed = True
                    except Exception as e:
                        print(f"[embedder] batch encode error for {file_path.name}: {e}", file=sys.stderr)
                        vecs = [[0.0] * vector_dim] * len(page_texts)
                        embedder_failed = True
                else:
                    vecs = [[0.0] * vector_dim] * len(page_texts)
            else:
                vecs = [[0.0] * vector_dim] * len(page_texts)

            zero_vec = [0.0] * vector_dim
            zero_vec_str = f"[{', '.join(str(x) for x in zero_vec)}]"
            title_esc = sql_escape(title)
            fpath_esc = sql_escape(fpath)
            fname_esc = sql_escape(file_path.name)
            content_esc = sql_escape(content)
            parent_sql = (f"INSERT INTO {safe_table} (embedding, filename, content, page, file_type, "
                         f"parent_doc_id, doc_path, chunk_index, total_chunks, title) "
                         f"VALUES ({zero_vec_str}, '{fname_esc}', '{content_esc}', "
                         f"0, '{ext[1:]}', NULL, '{fpath_esc}', 0, {total_chunks}, '{title_esc}')")
            result = db_client.execute(parent_sql)
            if "Error" in result:
                raise Exception(f"INSERT failed: {result}")
            parent_id = parse_id_from_result(result)
            if parent_id is None:
                parent_id = -1

            for chunk_idx, page_data in enumerate(pages):
                chunk_vec = vecs[chunk_idx] if chunk_idx < len(vecs) else [0.0] * vector_dim
                chunk_vec_str = f"[{', '.join(str(x) for x in chunk_vec)}]"
                chunk_content = page_data.get("content", "")
                page_num = page_data.get("page", 0)
                chunk_esc = sql_escape(chunk_content)
                child_sql = (f"INSERT INTO {safe_table} (embedding, filename, content, page, file_type, "
                             f"parent_doc_id, doc_path, chunk_index, total_chunks, title) "
                             f"VALUES ({chunk_vec_str}, '{fname_esc}', '{chunk_esc}', "
                             f"{page_num}, '{ext[1:]}', {parent_id}, '{fpath_esc}', {chunk_idx + 1}, {total_chunks}, '{title_esc}')")
                child_result = db_client.execute(child_sql)
                if "Error" in child_result:
                    raise Exception(f"INSERT failed: {child_result}")

            log_import(safe_table, fpath, file_path.name, fsize, fhash, chash, 1, total_chunks, "ok" if not embedder_failed else "embedder_failed")
            stats["imported"] += 1
            imported_files.append(file_path.name)
            if embedder_failed:
                error_details.append(f"{file_path.name}: Embedding generation failed, stored zero vectors")

        except Exception as e:
            stats["errors"] += 1
            error_details.append(f"{file_path.name}: {str(e)}")
            try:
                log_import(safe_table, fpath, file_path.name, fsize, fhash, chash, 0, 0, "error")
            except Exception:
                pass

        if (i + 1) % 5 == 0:
            print(f"[progress] Processed {i + 1}/{total_files} files...", file=sys.stderr, flush=True)

    _cleanup_tmp_dir(tmp_uuid_dir)
    tmp_uuid_dir = None

    embedder_info = f"yes ({EMBEDDER_MODEL})" if HAS_EMBEDDER else "no (install sentence-transformers)"
    summary = [
        f"Import completed:",
        f"  Files processed: {total_files}",
        f"  Imported: {stats['imported']}",
        f"  Skipped (already imported or too large): {stats['skipped']}",
        f"  Errors: {stats['errors']}",
        f"  Embeddings: {embedder_info}",
        "",
    ]
    if imported_files:
        summary.append("Imported files:")
        for fn in imported_files[:20]:
            summary.append(f"  - {fn}")
        if len(imported_files) > 20:
            summary.append(f"  ... and {len(imported_files) - 20} more")
    if error_details:
        summary.append("")
        summary.append("Issues:")
        for err in error_details[:20]:
            summary.append(f"  ! {err}")
    return {"content": [{"type": "text", "text": "\n".join(summary)}]}


async def handle_ingest_chunked(args: dict[str, Any]) -> dict[str, Any]:
    table = args.get("table")
    file_path = args.get("file_path")
    chunk_size = args.get("chunk_size", 500)
    overlap = args.get("overlap", 50)
    vector_dim = args.get("vector_dim", DEFAULT_VECTOR_DIM)
    tmp_uuid_dir = None

    if not table or not file_path:
        return {"content": [{"type": "text", "text": "Error: table and file_path are required"}], "isError": True}

    try:
        safe_table = sql_safe_identifier(table)
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid table name: {e}"}], "isError": True}

    path = Path(file_path)
    if not path.exists():
        return {"content": [{"type": "text", "text": f"Error: File not found: {file_path}"}], "isError": True}

    fsize = path.stat().st_size
    if fsize > MAX_FILE_SIZE_MB * 1024 * 1024:
        return {"content": [{"type": "text", "text": f"Error: File too large ({fsize / 1024 / 1024:.1f} MB > {MAX_FILE_SIZE_MB} MB limit)"}], "isError": True}

    if not HAS_EMBEDDER:
        return {"content": [{"type": "text", "text": "Error: sentence-transformers not installed. Install with: pip install sentence-transformers"}], "isError": True}

    embedder = get_embedder()
    if not embedder:
        return {"content": [{"type": "text", "text": "Error: Failed to load embedder model"}], "isError": True}

    try:
        tmp_uuid_dir = _create_tmp_uuid_dir()
        md_path, ext = _convert_to_markdown(file_path, tmp_uuid_dir)
        parsed = parse_md(str(md_path))
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Error: {e}"}], "isError": True}
    except ImportError as e:
        return {"content": [{"type": "text", "text": f"Error: Missing dependency for {ext} files: {e}"}], "isError": True}
    except Exception as e:
        error_msg = f"Error converting file to markdown: {e}"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        return {"content": [{"type": "text", "text": error_msg}], "isError": True}

    full_text = parsed.get("content", "")
    if not full_text.strip():
        error_msg = "Error: No text content extracted from file"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        return {"content": [{"type": "text", "text": error_msg}], "isError": True}

    chunks = smart_chunk(full_text, target_chars=chunk_size, overlap=overlap)
    if not chunks:
        error_msg = "Error: No chunks generated from text"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        return {"content": [{"type": "text", "text": error_msg}], "isError": True}

    total_chunks = len(chunks)
    source_file = str(path.resolve())

    db_client.execute(f"CREATE TABLE IF NOT EXISTS {safe_table} (embedding VECTOR({vector_dim}), content TEXT, source_file TEXT, chunk_index INT, total_chunks INT, chunk_hash TEXT)")

    existing_hashes = set()
    try:
        result = db_client.execute(f"SELECT chunk_hash FROM {safe_table} WHERE source_file = '{sql_escape(source_file)}'")
        for line in result.split("\n"):
            m = re.search(r"chunk_hash='([^']+)'", line)
            if m:
                existing_hashes.add(m.group(1))
    except Exception:
        pass

    chunks_to_embed = []
    chunks_to_skip = []
    for i, chunk in enumerate(chunks):
        h = compute_chunk_hash(chunk)
        if h in existing_hashes:
            chunks_to_skip.append(i)
        else:
            chunks_to_embed.append((i, chunk, h))

    if not chunks_to_embed:
        _cleanup_tmp_dir(tmp_uuid_dir)
        return {"content": [{"type": "text", "text": f"Ingest complete (all {total_chunks} chunks already exist)"}]}

    texts_to_embed = [c[1] for c in chunks_to_embed]
    chunk_indices = [c[0] for c in chunks_to_embed]
    chunk_hashes = [c[2] for c in chunks_to_embed]

    try:
        vecs = embedder.encode(texts_to_embed, convert_to_numpy=True, normalize_embeddings=True)
        vecs = vecs.tolist() if hasattr(vecs, 'tolist') else list(vecs)
        if len(vecs) != len(texts_to_embed) or (vecs and len(vecs[0]) != vector_dim):
            error_msg = f"Error: Embedding dimension mismatch (expected {vector_dim})"
            if tmp_uuid_dir:
                error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
            return {"content": [{"type": "text", "text": error_msg}], "isError": True}
    except Exception as e:
        error_msg = f"Error generating embeddings: {e}"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        return {"content": [{"type": "text", "text": error_msg}], "isError": True}

    inserted = 0
    try:
        for batch_start in range(0, len(chunks_to_embed), 100):
            batch_end = min(batch_start + 100, len(chunks_to_embed))
            for j in range(batch_start, batch_end):
                vec = vecs[j - batch_start]
                idx = chunk_indices[j]
                h = chunk_hashes[j]
                chunk_text = chunks[idx]
                vec_str = f"[{', '.join(str(x) for x in vec)}]"
                content_esc = sql_escape(chunk_text)
                sql = (f"INSERT INTO {safe_table} (embedding, content, source_file, chunk_index, total_chunks, chunk_hash) "
                       f"VALUES ({vec_str}, '{content_esc}', '{sql_escape(source_file)}', {idx}, {total_chunks}, '{h}')")
                result = db_client.execute(sql)
                if "Error" not in result:
                    inserted += 1

            print(f"  Processed {batch_end}/{len(chunks_to_embed)} chunks...", file=sys.stderr, flush=True)

        _cleanup_tmp_dir(tmp_uuid_dir)
        return {"content": [{"type": "text", "text": f"Ingest complete. Inserted {inserted} new chunks, skipped {len(chunks_to_skip)} duplicates (of {total_chunks} total chunks)."}]}
    except Exception as e:
        error_msg = f"Error inserting to database: {e}"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        return {"content": [{"type": "text", "text": error_msg}], "isError": True}


async def handle_ingest_joplin(args: dict[str, Any]) -> dict[str, Any]:
    table = args.get("table")
    note_id = args.get("note_id")
    note_content = args.get("note_content")
    note_title = args.get("note_title", "")
    note_tags = args.get("note_tags", "")
    created_time = args.get("created_time", 0)
    updated_time = args.get("updated_time", 0)
    chunk_size = args.get("chunk_size", 500)
    overlap = args.get("overlap", 50)
    vector_dim = args.get("vector_dim", DEFAULT_VECTOR_DIM)

    if not table or not note_id:
        return {"content": [{"type": "text", "text": "Error: table and note_id are required"}], "isError": True}
    if not note_content:
        return {"content": [{"type": "text", "text": "Error: note_content is required"}], "isError": True}

    try:
        safe_table = sql_safe_identifier(table)
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid table name: {e}"}], "isError": True}

    if not HAS_EMBEDDER:
        return {"content": [{"type": "text", "text": "Error: sentence-transformers not installed."}], "isError": True}

    embedder = get_embedder()
    if not embedder:
        return {"content": [{"type": "text", "text": "Error: Failed to load embedder"}], "isError": True}

    note_content = note_content.strip()
    if not note_content:
        return {"content": [{"type": "text", "text": "Error: Note content is empty"}], "isError": True}

    chunks = smart_chunk(note_content, target_chars=chunk_size, overlap=overlap)
    total_chunks = len(chunks)
    if not chunks:
        return {"content": [{"type": "text", "text": "Error: No chunks generated from note content"}], "isError": True}

    source_file = f"joplin:{note_id}"

    db_client.execute(f"CREATE TABLE IF NOT EXISTS {safe_table} (embedding VECTOR({vector_dim}), content TEXT, source_file TEXT, note_title TEXT, note_tags TEXT, created_time INT, updated_time INT, chunk_index INT, total_chunks INT, chunk_hash TEXT)")

    existing_hashes = set()
    try:
        result = db_client.execute(f"SELECT chunk_hash FROM {safe_table} WHERE source_file = '{sql_escape(source_file)}'")
        for line in result.split("\n"):
            m = re.search(r"chunk_hash='([^']+)'", line)
            if m:
                existing_hashes.add(m.group(1))
    except Exception:
        pass

    chunks_to_embed = []
    chunks_to_skip = []
    for i, chunk in enumerate(chunks):
        h = compute_chunk_hash(chunk)
        if h in existing_hashes:
            chunks_to_skip.append(i)
        else:
            chunks_to_embed.append((i, chunk, h))

    if not chunks_to_embed:
        return {"content": [{"type": "text", "text": f"Ingest complete (all {total_chunks} chunks already exist from '{note_title or note_id}')"}]}

    texts_to_embed = [c[1] for c in chunks_to_embed]
    chunk_indices = [c[0] for c in chunks_to_embed]
    chunk_hashes = [c[2] for c in chunks_to_embed]

    try:
        vecs = embedder.encode(texts_to_embed, convert_to_numpy=True, normalize_embeddings=True)
        vecs = vecs.tolist() if hasattr(vecs, 'tolist') else list(vecs)
        if len(vecs) != len(texts_to_embed) or (vecs and len(vecs[0]) != vector_dim):
            return {"content": [{"type": "text", "text": f"Error: Embedding dimension mismatch (expected {vector_dim})"}], "isError": True}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error generating embeddings: {e}"}], "isError": True}

    inserted = 0
    title_esc = sql_escape(note_title)
    tags_esc = sql_escape(note_tags)

    for batch_start in range(0, len(chunks_to_embed), 100):
        batch_end = min(batch_start + 100, len(chunks_to_embed))
        for j in range(batch_start, batch_end):
            vec = vecs[j - batch_start]
            idx = chunk_indices[j]
            h = chunk_hashes[j]
            chunk_text = chunks[idx]
            vec_str = f"[{', '.join(str(x) for x in vec)}]"
            content_esc = sql_escape(chunk_text)
            sql = (f"INSERT INTO {safe_table} (embedding, content, source_file, note_title, note_tags, created_time, updated_time, chunk_index, total_chunks, chunk_hash) "
                   f"VALUES ({vec_str}, '{content_esc}', '{sql_escape(source_file)}', '{title_esc}', '{tags_esc}', {created_time}, {updated_time}, {idx}, {total_chunks}, '{h}')")
            result = db_client.execute(sql)
            if "Error" not in result:
                inserted += 1

        print(f"  Processed {batch_end}/{len(chunks_to_embed)} chunks...", file=sys.stderr, flush=True)

    return {"content": [{"type": "text", "text": f"Ingest complete. Inserted {inserted} new chunks, skipped {len(chunks_to_skip)} duplicates (of {total_chunks} total chunks from '{note_title or note_id}')."}]}


# ==================== Job Tracking for Async Ingest ====================

_jobs: dict[str, dict] = {}
_job_counter = 0

def _next_job_id() -> str:
    global _job_counter
    _job_counter += 1
    return f"job_{_job_counter:06d}"


def _start_ingest_job(file_path: str, table: str, chunk_size: int, overlap: int, vector_dim: int):
    """Start async ingest job in background thread."""
    import threading

    job_id = _next_job_id()
    job_info = {
        "id": job_id,
        "file_path": file_path,
        "table": table,
        "status": "processing",
        "total_chunks": 0,
        "processed_chunks": 0,
        "inserted": 0,
        "skipped": 0,
        "error": None,
        "started_at": time.time(),
    }
    _jobs[job_id] = job_info

    def process():
        try:
            _process_ingest_job(job_id)
        except Exception as e:
            _jobs[job_id]["status"] = "failed"
            _jobs[job_id]["error"] = str(e)

    thread = threading.Thread(target=process, daemon=True)
    thread.start()
    return job_id


def _process_ingest_job(job_id: str):
    """Process ingest job in background."""
    job = _jobs[job_id]
    path = Path(job["file_path"])
    table = job["table"]
    tmp_uuid_dir = None

    try:
        safe_table = sql_safe_identifier(table)
    except ValueError:
        job["status"] = "failed"
        job["error"] = "Invalid table name"
        return

    try:
        tmp_uuid_dir = _create_tmp_uuid_dir()
        md_path, ext = _convert_to_markdown(str(path), tmp_uuid_dir)
        parsed = parse_md(str(md_path))
    except ValueError as e:
        job["status"] = "failed"
        job["error"] = str(e)
        return
    except Exception as e:
        job["status"] = "failed"
        error_msg = f"Convert to markdown error: {e}"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        job["error"] = error_msg
        return

    full_text = parsed.get("content", "")
    if not full_text.strip():
        job["status"] = "failed"
        job["error"] = "No text content"
        if tmp_uuid_dir:
            job["error"] += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        return

    chunk_size = job.get("chunk_size", 500)
    overlap = job.get("overlap", 50)
    vector_dim = job.get("vector_dim", 384)

    chunks = smart_chunk(full_text, target_chars=chunk_size, overlap=overlap)
    total_chunks = len(chunks)
    job["total_chunks"] = total_chunks

    if not chunks:
        job["status"] = "failed"
        job["error"] = "No chunks generated"
        if tmp_uuid_dir:
            job["error"] += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        return

    source_file = str(path.resolve())
    db_client.execute(f"CREATE TABLE IF NOT EXISTS {safe_table} (embedding VECTOR({vector_dim}), content TEXT, source_file TEXT, chunk_index INT, total_chunks INT, chunk_hash TEXT)")

    existing_hashes = set()
    try:
        result = db_client.execute(f"SELECT chunk_hash FROM {safe_table} WHERE source_file = '{sql_escape(source_file)}'")
        for line in result.split("\n"):
            m = re.search(r"chunk_hash='([^']+)'", line)
            if m:
                existing_hashes.add(m.group(1))
    except Exception:
        pass

    chunks_to_embed = []
    chunks_to_skip = []
    for i, chunk in enumerate(chunks):
        h = compute_chunk_hash(chunk)
        if h in existing_hashes:
            chunks_to_skip.append(i)
        else:
            chunks_to_embed.append((i, chunk, h))

    job["skipped"] = len(chunks_to_skip)

    if not chunks_to_embed:
        _cleanup_tmp_dir(tmp_uuid_dir)
        job["status"] = "completed"
        job["inserted"] = 0
        return

    embedder = get_embedder()
    if not embedder or not HAS_EMBEDDER:
        job["status"] = "failed"
        error_msg = "Embedder not available"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        job["error"] = error_msg
        return

    texts_to_embed = [c[1] for c in chunks_to_embed]
    chunk_indices = [c[0] for c in chunks_to_embed]
    chunk_hashes = [c[2] for c in chunks_to_embed]

    try:
        vecs = embedder.encode(texts_to_embed, convert_to_numpy=True, normalize_embeddings=True)
        vecs = vecs.tolist() if hasattr(vecs, 'tolist') else list(vecs)
    except Exception as e:
        job["status"] = "failed"
        error_msg = f"Embedding error: {e}"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        job["error"] = error_msg
        return

    BATCH_SIZE = 100
    inserted = 0

    try:
        for batch_start in range(0, len(chunks_to_embed), BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, len(chunks_to_embed))
            batch_vecs = vecs[batch_start:batch_end]
            batch_indices = chunk_indices[batch_start:batch_end]
            batch_hashes = chunk_hashes[batch_start:batch_end]

            values_parts = []
            for j in range(batch_end - batch_start):
                vec = batch_vecs[j]
                idx = batch_indices[j]
                h = batch_hashes[j]
                chunk_text = chunks[idx]
                vec_str = f"[{', '.join(str(x) for x in vec)}]"
                content_esc = sql_escape(chunk_text)
                values_parts.append(f"({vec_str}, '{content_esc}', '{sql_escape(source_file)}', {idx}, {total_chunks}, '{h}')")

            sql = f"INSERT INTO {safe_table} (embedding, content, source_file, chunk_index, total_chunks, chunk_hash) VALUES {', '.join(values_parts)}"
            result = db_client.execute(sql)
            if "Error" not in result:
                inserted += len(batch_vecs)

            job["processed_chunks"] = batch_end
            job["inserted"] = inserted

        _cleanup_tmp_dir(tmp_uuid_dir)
        job["status"] = "completed"
        job["inserted"] = inserted
    except Exception as e:
        job["status"] = "failed"
        error_msg = f"DB insert error: {e}"
        if tmp_uuid_dir:
            error_msg += f"\n[debug] tmp preserved at: {tmp_uuid_dir}"
        job["error"] = error_msg


async def handle_ingest_async(args: dict[str, Any]) -> dict[str, Any]:
    table = args.get("table")
    file_path = args.get("file_path")
    chunk_size = args.get("chunk_size", 500)
    overlap = args.get("overlap", 50)
    vector_dim = args.get("vector_dim", DEFAULT_VECTOR_DIM)

    if not table or not file_path:
        return {"content": [{"type": "text", "text": "Error: table and file_path are required"}], "isError": True}

    path = Path(file_path)
    if not path.exists():
        return {"content": [{"type": "text", "text": f"Error: File not found: {file_path}"}], "isError": True}

    fsize = path.stat().st_size
    if fsize > MAX_FILE_SIZE_MB * 1024 * 1024:
        return {"content": [{"type": "text", "text": f"Error: File too large ({fsize / 1024 / 1024:.1f} MB > {MAX_FILE_SIZE_MB} MB limit)"}], "isError": True}

    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return {"content": [{"type": "text", "text": f"Error: Unsupported file type '{ext}'. Supported: {', '.join(SUPPORTED_EXTENSIONS)}"}], "isError": True}

    job_id = _start_ingest_job(file_path, table, chunk_size, overlap, vector_dim)
    return {"content": [{"type": "text", "text": f"Job started: {job_id}. Poll with pardusdb_ingest_status with job_id='{job_id}'"}]}


async def handle_ingest_status(args: dict[str, Any]) -> dict[str, Any]:
    job_id = args.get("job_id")
    if not job_id:
        return {"content": [{"type": "text", "text": "Error: job_id is required"}], "isError": True}

    if job_id not in _jobs:
        return {"content": [{"type": "text", "text": f"Job not found: {job_id}"}], "isError": True}

    job = _jobs[job_id]
    elapsed = time.time() - job["started_at"]
    progress = f"{job['processed_chunks']}/{job['total_chunks']}" if job['total_chunks'] > 0 else "0/0"

    result_text = f"""Job: {job_id}
Status: {job['status']}
File: {job['file_path']}
Table: {job['table']}
Progress: {progress}
Inserted: {job['inserted']}
Skipped (duplicates): {job['skipped']}
Elapsed: {elapsed:.1f}s"""
    if job.get("error"):
        result_text += f"\nError: {job['error']}"

    return {"content": [{"type": "text", "text": result_text}]}


# ==================== Health Check Tool ====================

async def handle_health_check(args: dict[str, Any]) -> dict[str, Any]:
    target_table = args.get("table")
    repair = args.get("repair", False)

    db_path = db_client.get_db_path()
    if not db_path:
        return {"content": [{"type": "text", "text": "Error: No database opened. Use open_database first."}], "isError": True}

    if not Path(db_path).exists():
        return {"content": [{"type": "text", "text": f"Error: Database file not found: {db_path}"}], "isError": True}

    if target_table:
        try:
            target_table = sql_safe_identifier(target_table)
        except ValueError as e:
            return {"content": [{"type": "text", "text": f"Invalid table name: {e}"}], "isError": True}

    db_size = os.path.getsize(db_path)
    report = [f"Database Health Report", f"Database: {db_path}", f"Size: {db_size / 1024:.2f} KB", ""]

    try:
        tables_result = db_client.execute("SHOW TABLES")
        table_names = re.findall(r"Table '(\w+)'", tables_result)
        if not table_names:
            report.append("No tables found.")
            return {"content": [{"type": "text", "text": "\n".join(report)}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error listing tables: {e}"}], "isError": True}

    tables_to_check = [target_table] if target_table else [t for t in table_names if t != "__import_log__"]

    for tbl in tables_to_check:
        safe_tbl = sql_safe_identifier(tbl)
        if tbl not in table_names:
            report.append(f"Table: {tbl}")
            report.append("  ❌ Table does not exist")
            report.append("")
            continue

        report.append(f"Table: {tbl}")
        has_warnings = False

        try:
            count_result = db_client.execute(f"SELECT COUNT(*) FROM {safe_tbl}")
            total_rows = parse_count_from_result(count_result)
            report.append(f"  Total records: {total_rows}")
        except Exception as e:
            report.append(f"  ⚠️  Could not count rows: {e}")

        try:
            orphans_result = db_client.execute(
                f"SELECT COUNT(*) FROM {safe_tbl} WHERE parent_doc_id IS NOT NULL "
                f"AND parent_doc_id NOT IN (SELECT id FROM {safe_tbl} WHERE parent_doc_id IS NULL)"
            )
            orphan_count = parse_count_from_result(orphans_result)
            if orphan_count > 0:
                report.append(f"  ❌ Orphan children found: {orphan_count}")
                has_warnings = True
            else:
                report.append(f"  ✅ No orphan children")
        except Exception as e:
            report.append(f"  ⚠️  Could not check orphans: {e}")

        try:
            dup_result = db_client.execute(
                f"SELECT doc_path, COUNT(*) as cnt FROM {safe_tbl} "
                f"WHERE parent_doc_id IS NULL AND doc_path IS NOT NULL "
                f"GROUP BY doc_path HAVING cnt > 1"
            )
            if "id=" in dup_result or dup_result.strip():
                report.append(f"  ⚠️  Duplicate parent documents found")
                has_warnings = True
            else:
                report.append(f"  ✅ No duplicate parent documents")
        except Exception:
            pass

        try:
            zero_result = db_client.execute(
                f"SELECT COUNT(*) FROM {safe_tbl} WHERE embedding = '[{', '.join(['0.0'] * DEFAULT_VECTOR_DIM)}]'"
            )
            zero_count = parse_count_from_result(zero_result)
            if zero_count > 0:
                report.append(f"  ⚠️  Records with zero vectors: {zero_count}")
                has_warnings = True
            else:
                report.append(f"  ✅ All records have valid embeddings")
        except Exception:
            pass

        if not has_warnings:
            report.append(f"  ✅ Health: PASS")
        else:
            report.append(f"  ⚠️  Health: WARNINGS")
        report.append("")

    try:
        if "__import_log__" in table_names:
            log_count = db_client.execute("SELECT COUNT(*) FROM __import_log__")
            log_rows = parse_count_from_result(log_count)
            report.append(f"Import Log: {log_rows} entries")
            ok_count_result = db_client.execute("SELECT COUNT(*) FROM __import_log__ WHERE status = 'ok'")
            ok_rows = parse_count_from_result(ok_count_result)
            report.append(f"  Successful imports: {ok_rows}")
        else:
            report.append("Import Log: not found")
    except Exception as e:
        report.append(f"⚠️  Could not check import log: {e}")

    return {"content": [{"type": "text", "text": "\n".join(report)}]}


# ==================== Schema Tool ====================

async def handle_get_schema(args: dict[str, Any]) -> dict[str, Any]:
    table = args.get("table")
    if not table:
        return {"content": [{"type": "text", "text": "Error: Table name is required"}], "isError": True}
    try:
        safe_table = sql_safe_identifier(table)
    except ValueError as e:
        return {"content": [{"type": "text", "text": f"Invalid table name: {e}"}], "isError": True}
    try:
        tables_result = db_client.execute("SHOW TABLES")
        if table not in tables_result:
            return {"content": [{"type": "text", "text": f"Table '{table}' does not exist"}], "isError": True}

        count_result = db_client.execute(f"SELECT COUNT(*) FROM {safe_table}")
        row_count = parse_count_from_result(count_result)

        output = [f"Schema for table: {table}", f"Total rows: {row_count}", ""]

        output.append("Note: Use raw SQL to inspect column types, as SHOW TABLES does not expose schema details.")
        output.append("Known columns for import tables: embedding, filename, content, page, file_type, parent_doc_id, doc_path, chunk_index, total_chunks, title")

        return {"content": [{"type": "text", "text": "\n".join(output)}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error getting schema: {e}"}], "isError": True}


# ==================== Import Status Tool ====================

async def handle_import_status(args: dict[str, Any]) -> dict[str, Any]:
    action = args.get("action", "list")
    table = args.get("table")
    doc_path = args.get("doc_path")

    try:
        tables_result = db_client.execute("SHOW TABLES")
        if "__import_log__" not in tables_result:
            return {"content": [{"type": "text", "text": "No import history found (import_log table does not exist)"}]}

        if action == "list":
            where_clauses = []
            if table:
                where_clauses.append(f"table_name = '{sql_escape(table)}'")
            if doc_path:
                where_clauses.append(f"doc_path = '{sql_escape(doc_path)}'")
            where = " AND ".join(where_clauses) if where_clauses else "1=1"
            sql = f"SELECT filename, table_name, file_size, imported_at, total_parents, total_children, status FROM __import_log__ WHERE {where} ORDER BY imported_at DESC LIMIT 50"
            result = db_client.execute(sql)
            return {"content": [{"type": "text", "text": f"Import History:\n\n{result}"}]}

        elif action == "reset":
            if not table and not doc_path:
                return {"content": [{"type": "text", "text": "Error: Provide table or doc_path to reset"}], "isError": True}
            where_clauses = []
            if table:
                where_clauses.append(f"table_name = '{sql_escape(table)}'")
            if doc_path:
                where_clauses.append(f"doc_path = '{sql_escape(doc_path)}'")
            where = " AND ".join(where_clauses)
            sql = f"DELETE FROM __import_log__ WHERE {where}"
            db_client.execute(sql)
            return {"content": [{"type": "text", "text": f"Import records reset for: {table or ''} {doc_path or ''}"}]}

        else:
            return {"content": [{"type": "text", "text": f"Unknown action: {action}. Use 'list' or 'reset'."}], "isError": True}

    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error: {e}"}], "isError": True}


async def handle_get_stats(args: dict[str, Any]) -> dict[str, Any]:
    """Get token savings statistics."""
    try:
        stats = get_token_stats()
        config = stats["config"]
        session = stats["session"]
        total = stats["total"]

        response_text = f"""Token Savings Dashboard
═══════════════════════════════════════════════════════

CURRENT MODEL: {config['model']}
Provider: {config['provider']}
Context Window: {config['context_window']:,} tokens

─────────────── SESSION ───────────────
Queries: {session['queries']}
Tokens sent: {session['tokens_sent']:,}
Tokens if full doc: {session['tokens_if_full']:,}
Savings: {session['savings']:,} tokens ({session['savings_percent']}%)
Chunks returned: {session['chunks_returned']}

─────────────── TOTAL ───────────────
Total Queries: {total['queries']}
Total Tokens sent: {total['tokens_sent']:,}
Total Tokens if full: {total['tokens_if_full']:,}
Total Savings: {total['savings']:,} tokens ({total['savings_percent']}%)
Total Chunks: {total['chunks_returned']}
"""
        return {"content": [{"type": "text", "text": response_text}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error getting stats: {e}"}], "isError": True}


async def handle_set_model(args: dict[str, Any]) -> dict[str, Any]:
    """Set the current LLM model for accurate token tracking."""
    model = args.get("model")
    if not model:
        return {"content": [{"type": "text", "text": "Error: model name is required"}], "isError": True}
    try:
        result = set_current_model(model)
        return {"content": [{"type": "text", "text": f"Model updated: {result['model']}\nContext window: {result['context_window']:,} tokens\nProvider: {result['provider']}"}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error setting model: {e}"}], "isError": True}


async def handle_reset_stats(args: dict[str, Any]) -> dict[str, Any]:
    """Reset session statistics."""
    try:
        result = reset_session_stats()
        return {"content": [{"type": "text", "text": f"Session stats reset successfully"}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error resetting stats: {e}"}], "isError": True}


# ==================== Tool Definitions ====================

TOOLS = [
    Tool(
        name="pardusdb_create_database",
        description="Create a new PardusDB database file at the specified path",
        inputSchema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path for the new .pardus database file"},
            },
            "required": ["path"],
        },
    ),
    Tool(
        name="pardusdb_open_database",
        description="Open an existing PardusDB database file",
        inputSchema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the existing .pardus database file"},
            },
            "required": ["path"],
        },
    ),
    Tool(
        name="pardusdb_create_table",
        description="Create a new table for storing vectors with optional metadata columns",
        inputSchema={
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Name of the table"},
                "vector_dim": {"type": "number", "description": "Dimension of the vectors (e.g. 768)"},
                "metadata_schema": {
                    "type": "object",
                    "description": "Optional metadata columns: {column_name: type}. Types: str, int, float, bool",
                    "additionalProperties": {"type": "string"},
                },
            },
            "required": ["name", "vector_dim"],
        },
    ),
    Tool(
        name="pardusdb_insert_vector",
        description="Insert a single vector with optional metadata into a table",
        inputSchema={
            "type": "object",
            "properties": {
                "vector": {"type": "array", "items": {"type": "number"}, "description": "The embedding vector"},
                "metadata": {"type": "object", "description": "Optional metadata to store with the vector"},
                "table": {"type": "string", "description": "Table name (uses current table if not specified)"},
            },
            "required": ["vector"],
        },
    ),
    Tool(
        name="pardusdb_batch_insert",
        description="Insert multiple vectors efficiently in a batch",
        inputSchema={
            "type": "object",
            "properties": {
                "vectors": {
                    "type": "array",
                    "items": {"type": "array", "items": {"type": "number"}},
                    "description": "Array of embedding vectors",
                },
                "metadata_list": {"type": "array", "items": {"type": "object"}, "description": "Optional metadata per vector"},
                "table": {"type": "string", "description": "Table name (uses current table if not specified)"},
            },
            "required": ["vectors"],
        },
    ),
    Tool(
        name="pardusdb_search_similar",
        description="Search for vectors similar to a query vector using cosine similarity",
        inputSchema={
            "type": "object",
            "properties": {
                "query_vector": {"type": "array", "items": {"type": "number"}, "description": "The query embedding vector"},
                "k": {"type": "number", "description": "Number of results (default: 10)"},
                "table": {"type": "string", "description": "Table name (uses current table if not specified)"},
            },
            "required": ["query_vector"],
        },
    ),
    Tool(
        name="pardusdb_search_text",
        description="Search for vectors similar to a text query using semantic similarity. Generates an embedding from the text and searches the database.",
        inputSchema={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "The text query to search for"},
                "k": {"type": "number", "description": "Number of results (default: 10)"},
                "table": {"type": "string", "description": "Table name (uses current table if not specified)"},
                "vector_dim": {"type": "number", "description": f"Embedding dimension (default: {DEFAULT_VECTOR_DIM})"},
            },
            "required": ["query"],
        },
    ),
    Tool(
        name="pardusdb_execute_sql",
        description="Execute raw SQL commands on the database",
        inputSchema={
            "type": "object",
            "properties": {
                "sql": {"type": "string", "description": "SQL command to execute"},
            },
            "required": ["sql"],
        },
    ),
    Tool(
        name="pardusdb_list_tables",
        description="List all tables in the current database",
        inputSchema={"type": "object", "properties": {}},
    ),
    Tool(
        name="pardusdb_use_table",
        description="Set the current table for subsequent operations",
        inputSchema={
            "type": "object",
            "properties": {
                "table": {"type": "string", "description": "Name of the table to use"},
            },
            "required": ["table"],
        },
    ),
    Tool(
        name="pardusdb_status",
        description="Get the current status of the database connection",
        inputSchema={"type": "object", "properties": {}},
    ),
    Tool(
        name="pardusdb_import_text",
        description="Import documents from a directory into a table. Scans recursively and imports PDF, CSV, DOCX, XLSX, JSON, JSONL, MD, and TXT files with automatic embeddings. Multi-page files create parent + child records with parent-child tracking. Skips already-imported files (by SHA256 hash).",
        inputSchema={
            "type": "object",
            "properties": {
                "dir_path": {"type": "string", "description": "Directory path to scan for documents"},
                "table": {"type": "string", "description": "Target table name (created if not exists)"},
                "file_patterns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "File extensions to include. Default: all supported (.pdf, .csv, .docx, .xlsx, .json, .jsonl, .md, .txt)",
                },
                "recursive": {"type": "boolean", "description": "Scan subdirectories recursively (default: true)"},
                "max_file_size_mb": {"type": "number", "description": f"Max file size in MB (default: {MAX_FILE_SIZE_MB})"},
                "vector_dim": {"type": "number", "description": f"Embedding dimension (default: {DEFAULT_VECTOR_DIM})"},
            },
            "required": ["dir_path", "table"],
        },
    ),
    Tool(
        name="pardusdb_ingest_chunked",
        description="Ingest a single document with smart sentence-aware chunking. Extracts full text, splits into coherent chunks of ~chunk_size characters at sentence boundaries, generates embeddings, and batch inserts. Skips duplicate chunks by content hash.",
        inputSchema={
            "type": "object",
            "properties": {
                "table": {"type": "string", "description": "Target table name"},
                "file_path": {"type": "string", "description": "Path to document file (PDF, DOCX, TXT, MD, CSV, JSON, JSONL, XLSX, XLS)"},
                "chunk_size": {"type": "integer", "description": "Target characters per chunk (default: 500)"},
                "overlap": {"type": "integer", "description": "Character overlap between chunks for continuity (default: 50)"},
                "vector_dim": {"type": "integer", "description": f"Embedding dimension (default: {DEFAULT_VECTOR_DIM})"},
            },
            "required": ["table", "file_path"],
        },
    ),
    Tool(
        name="pardusdb_ingest_joplin",
        description="Ingest a Joplin note into PardusDB with smart sentence-aware chunking. Use after joplin_read_note to pass the fetched note content, title, tags, and timestamps. Reuses smart_chunk for semantic coherence and skips duplicate chunks by content hash.",
        inputSchema={
            "type": "object",
            "properties": {
                "table": {"type": "string", "description": "Target table name"},
                "note_id": {"type": "string", "description": "Joplin note ID (used for source tracking and deduplication)"},
                "note_content": {"type": "string", "description": "Full note body content from joplin_read_note"},
                "note_title": {"type": "string", "description": "Note title from joplin_read_note"},
                "note_tags": {"type": "string", "description": "Comma-separated tags from joplin_read_note"},
                "created_time": {"type": "integer", "description": "Unix timestamp in ms from joplin_read_note (optional)"},
                "updated_time": {"type": "integer", "description": "Unix timestamp in ms from joplin_read_note (optional)"},
                "chunk_size": {"type": "integer", "description": "Target characters per chunk (default: 500)"},
                "overlap": {"type": "integer", "description": "Character overlap between chunks for continuity (default: 50)"},
                "vector_dim": {"type": "integer", "description": f"Embedding dimension (default: {DEFAULT_VECTOR_DIM})"},
            },
            "required": ["table", "note_id", "note_content", "note_title"],
        },
    ),
    Tool(
        name="pardusdb_ingest_async",
        description="Ingest a document asynchronously to avoid timeout. Starts processing immediately and returns a job_id for tracking. Poll with pardusdb_ingest_status to monitor progress. For large PDFs (50MB+), use this instead of ingest_chunked.",
        inputSchema={
            "type": "object",
            "properties": {
                "table": {"type": "string", "description": "Target table name"},
                "file_path": {"type": "string", "description": "Path to document file (PDF, DOCX, TXT, MD, CSV, JSON, JSONL, XLSX, XLS)"},
                "chunk_size": {"type": "integer", "description": "Target characters per chunk (default: 500)"},
                "overlap": {"type": "integer", "description": "Character overlap between chunks for continuity (default: 50)"},
                "vector_dim": {"type": "integer", "description": f"Embedding dimension (default: {DEFAULT_VECTOR_DIM})"},
            },
            "required": ["table", "file_path"],
        },
    ),
    Tool(
        name="pardusdb_ingest_status",
        description="Check the status of an async ingest job started with pardusdb_ingest_async. Returns progress, chunks processed, and any errors.",
        inputSchema={
            "type": "object",
            "properties": {
                "job_id": {"type": "string", "description": "Job ID returned by pardusdb_ingest_async"},
            },
            "required": ["job_id"],
        },
    ),
    Tool(
        name="pardusdb_health_check",
        description="Run integrity checks on the database. Verifies tables, detects orphan children, checks for duplicate parent documents, and validates embedding consistency.",
        inputSchema={
            "type": "object",
            "properties": {
                "table": {"type": "string", "description": "Specific table to check (checks all if omitted)"},
                "repair": {"type": "boolean", "description": "Attempt to fix issues (default: false)"},
            },
        },
    ),
    Tool(
        name="pardusdb_get_schema",
        description="Show the schema and statistics of a table",
        inputSchema={
            "type": "object",
            "properties": {
                "table": {"type": "string", "description": "Table name"},
            },
            "required": ["table"],
        },
    ),
    Tool(
        name="pardusdb_import_status",
        description="View or reset import history. Use 'list' to see imports, 'reset' to clear records for re-importing.",
        inputSchema={
            "type": "object",
            "properties": {
                "action": {"type": "string", "enum": ["list", "reset"], "description": "'list' to view history, 'reset' to clear records"},
                "table": {"type": "string", "description": "Filter by table name (optional)"},
                "doc_path": {"type": "string", "description": "Filter by file path (optional, for reset)"},
            },
            "required": ["action"],
        },
    ),
    Tool(
        name="pardusdb_get_stats",
        description="Get token savings statistics. Shows session and total stats including tokens sent, tokens if full doc, and savings percentage.",
        inputSchema={
            "type": "object",
            "properties": {},
        },
    ),
    Tool(
        name="pardusdb_set_model",
        description="Set the current LLM model for accurate token tracking. Updates the context window and provider info.",
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string", "description": "Model name (e.g., 'gpt-4o', 'claude-3-5-sonnet-20241022')"},
            },
            "required": ["model"],
        },
    ),
    Tool(
        name="pardusdb_reset_stats",
        description="Reset session statistics. Keeps total accumulated stats but starts a new session counter.",
        inputSchema={
            "type": "object",
            "properties": {},
        },
    ),
]


# ==================== Server Setup ====================

server = Server("pardusdb-mcp", "0.4.21")


@server.list_tools()
async def list_tools() -> list[Tool]:
    return TOOLS


@server.call_tool()
async def call_tool(name: str, args: dict[str, Any]) -> list[TextContent]:
    result: dict[str, Any]

    if name == "pardusdb_create_database":
        result = await handle_create_database(args)
    elif name == "pardusdb_open_database":
        result = await handle_open_database(args)
    elif name == "pardusdb_create_table":
        result = await handle_create_table(args)
    elif name == "pardusdb_insert_vector":
        result = await handle_insert_vector(args)
    elif name == "pardusdb_batch_insert":
        result = await handle_batch_insert(args)
    elif name == "pardusdb_search_similar":
        result = await handle_search_similar(args)
    elif name == "pardusdb_search_text":
        result = await handle_search_text(args)
    elif name == "pardusdb_execute_sql":
        result = await handle_execute_sql(args)
    elif name == "pardusdb_list_tables":
        result = await handle_list_tables()
    elif name == "pardusdb_use_table":
        result = await handle_use_table(args)
    elif name == "pardusdb_status":
        result = await handle_get_status()
    elif name == "pardusdb_import_text":
        result = await handle_import_text(args)
    elif name == "pardusdb_health_check":
        result = await handle_health_check(args)
    elif name == "pardusdb_get_schema":
        result = await handle_get_schema(args)
    elif name == "pardusdb_import_status":
        result = await handle_import_status(args)
    elif name == "pardusdb_ingest_chunked":
        result = await handle_ingest_chunked(args)
    elif name == "pardusdb_ingest_joplin":
        result = await handle_ingest_joplin(args)
    elif name == "pardusdb_ingest_async":
        result = await handle_ingest_async(args)
    elif name == "pardusdb_ingest_status":
        result = await handle_ingest_status(args)
    elif name == "pardusdb_get_stats":
        result = await handle_get_stats(args)
    elif name == "pardusdb_set_model":
        result = await handle_set_model(args)
    elif name == "pardusdb_reset_stats":
        result = await handle_reset_stats(args)
    else:
        result = {"content": [{"type": "text", "text": f"Unknown tool: {name}"}], "isError": True}

    is_error = result.pop("isError", False)
    return [TextContent(type="text", text=result["content"][0]["text"])]


async def main() -> None:
    async with stdio_server() as (read_stream, write_stream):
        await server.run(read_stream, write_stream, server.create_initialization_options())


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        sys.exit(0)
    except Exception as e:
        print(f"Fatal error: {e}", file=sys.stderr)
        sys.exit(1)
